"""Spreadsheet and CSV extraction."""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import openpyxl

from src.ingestion.extractors.common import BaseExtractor, ExtractionResult
from src.ingestion.models import Chunk, ConfidenceLevel, ExtractionMethod, Modality, TableMetadata
from src.utils.ids import generate_chunk_id


class XLSXExtractor(BaseExtractor):
    """Extract spreadsheet sheets into table-oriented chunks."""

    async def can_extract(self, file_path: Path) -> bool:
        """Return whether the file is tabular data we support."""

        return file_path.suffix.lower() in {".xlsx", ".xls", ".csv"}

    async def extract(self, file_path: Path, doc_id: str, folder_id: str) -> ExtractionResult:
        """Route extraction based on file type."""

        if file_path.suffix.lower() == ".csv":
            return await self._extract_csv(file_path, doc_id, folder_id)
        return await self._extract_xlsx(file_path, doc_id, folder_id)

    async def _extract_xlsx(self, file_path: Path, doc_id: str, folder_id: str) -> ExtractionResult:
        """Extract content from an XLSX workbook."""

        document = self._create_document_metadata(file_path, doc_id, folder_id, has_tables=True)
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            text_sections: list[str] = []
            chunks: list[Chunk] = []

            for table_index, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                rows = []
                for row in worksheet.iter_rows(values_only=True):
                    values = [str(cell).strip() if cell is not None else "" for cell in row]
                    if any(values):
                        rows.append(values)

                if not rows:
                    continue

                headers = rows[0]
                structured_rows = []
                for row in rows[1:101]:
                    structured_rows.append(
                        {
                            headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}": value
                            for index, value in enumerate(row)
                        }
                    )

                linearized = [f"Sheet: {sheet_name}", " | ".join(headers)]
                linearized.extend(" | ".join(row) for row in rows[1:51])
                table_text = "\n".join(linearized)
                text_sections.append(table_text)

                chunks.append(
                    Chunk(
                        chunk_id=generate_chunk_id(doc_id, table_index),
                        folder_id=folder_id,
                        source_doc_id=doc_id,
                        modality=Modality.TABLE,
                        content_text=table_text,
                        file_path=str(file_path.resolve()),
                        file_name=file_path.name,
                        chunk_index=table_index,
                        extraction_method=ExtractionMethod.TABLE_PARSE,
                        confidence=ConfidenceLevel.HIGH,
                        table_metadata=TableMetadata(
                            rows=max(len(rows) - 1, 0),
                            columns=len(headers),
                            headers=headers,
                            table_index=table_index,
                            structured_data=structured_rows,
                        ),
                        metadata={"sheet_name": sheet_name},
                        content_hash=hashlib.md5(table_text.encode("utf-8")).hexdigest(),
                    )
                )

            workbook.close()
            document.table_chunks = len(chunks)
            return ExtractionResult(document=document, text_content="\n\n".join(text_sections), chunks=chunks, success=True)
        except Exception as exc:
            return ExtractionResult(document=document, text_content="", success=False, error_message=str(exc))

    async def _extract_csv(self, file_path: Path, doc_id: str, folder_id: str) -> ExtractionResult:
        """Extract content from a CSV file."""

        document = self._create_document_metadata(file_path, doc_id, folder_id, has_tables=True)
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as handle:
                rows = list(csv.reader(handle))
            if not rows:
                return ExtractionResult(document=document, text_content="", success=True)

            headers = rows[0]
            data_rows = rows[1:]
            linearized = [" | ".join(headers)]
            linearized.extend(" | ".join(row) for row in data_rows[:50])
            table_text = "\n".join(linearized)

            chunk = Chunk(
                chunk_id=generate_chunk_id(doc_id, 0),
                folder_id=folder_id,
                source_doc_id=doc_id,
                modality=Modality.TABLE,
                content_text=table_text,
                file_path=str(file_path.resolve()),
                file_name=file_path.name,
                chunk_index=0,
                extraction_method=ExtractionMethod.TABLE_PARSE,
                confidence=ConfidenceLevel.HIGH,
                table_metadata=TableMetadata(
                    rows=len(data_rows),
                    columns=len(headers),
                    headers=headers,
                    table_index=0,
                    structured_data=[
                        {
                            headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}": value
                            for index, value in enumerate(row)
                        }
                        for row in data_rows[:100]
                    ],
                ),
                content_hash=hashlib.md5(table_text.encode("utf-8")).hexdigest(),
            )
            document.table_chunks = 1
            return ExtractionResult(document=document, text_content=table_text, chunks=[chunk], success=True)
        except Exception as exc:
            return ExtractionResult(document=document, text_content="", success=False, error_message=str(exc))
